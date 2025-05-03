from typing import List
from loguru import logger
import re
from openpyxl import load_workbook

from database import DatabaseManager
from models import PhoneWhiteList


class ExcelParsingError(Exception): ...


def parse_phone_number(phone: str | int) -> str | None:
    if isinstance(phone, int):
        phone = str(phone)
    if phone_match := re.match(r"[78]?(9\d*)", phone):
        return phone_match.group(1)
    else:
        return None


def get_phones_whitelist_from_xls(file) -> List[str]:
    wb = load_workbook(file)
    first_sheet = wb.worksheets[0]
    for column in first_sheet.iter_cols(max_col=first_sheet.max_column):
        if column[0].value == "Телефон":
            return [
                phone_num
                for cell in column[1:]
                if (phone_num := parse_phone_number(cell.value)) is not None
            ]
    raise ExcelParsingError("Cannot find phone column in the provided Excel file.")


async def add_phones_from_file(
    file, db_manager: DatabaseManager
) -> List[PhoneWhiteList]:
    numbers_in_file = get_phones_whitelist_from_xls(file)
    errors = []
    phones_added = []
    for phone_number in numbers_in_file:
        try:
            phones_added += await db_manager.add_phone(int(phone_number[-10:]))
        except Exception as e:
            errors.append(str(e))
    if errors:
        logger.warning(f"""{len(errors)} happend during parsing incoming phone list. \n
                       {"\n".join(set(errors))}
                       """)
    return phones_added
